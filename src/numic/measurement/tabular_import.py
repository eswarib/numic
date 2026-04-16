"""Parse CSV / Excel (.xlsx) rows into ``PatientMeasurementRecord`` parts."""

from __future__ import annotations

import csv
import re
from datetime import date, datetime
from io import BytesIO, StringIO
from typing import Any

from pydantic import ValidationError

from numic.api.schemas.measurement import (
    MeasurementContext,
    PatientInfo,
    TabularImportResponse,
    TabularImportRowOut,
    TabularImportSkippedRow,
)
from numic.api.schemas.scoring import VentricularMeasurements


def _norm_header(h: str) -> str:
    s = h.strip().lower()
    s = re.sub(r"\s+", "_", s)
    return s


_HEADER_TO_CANONICAL: dict[str, str] = {}
for canon, aliases in (
    (
        "vi_mm",
        ("vi_mm", "vi", "ventricular_index", "ventricular_index_mm", "vent_index_mm"),
    ),
    ("ahw_mm", ("ahw_mm", "ahw", "anterior_horn_width", "anterior_horn_width_mm")),
    ("tod_mm", ("tod_mm", "tod", "thalamo_occipital", "thalamo_occipital_distance_mm")),
    (
        "vi_percentile",
        ("vi_percentile", "percentile", "ga_vi_percentile", "vi_pct"),
    ),
    (
        "vi_p97_reference_mm",
        ("vi_p97_reference_mm", "p97_ref_mm", "vi_nomogram_mm", "p97_mm"),
    ),
    ("external_ref", ("external_ref", "patient_id", "mrn", "episode_id", "subject_id")),
    (
        "measured_at",
        ("measured_at", "measurement_datetime", "scan_datetime", "study_datetime", "acquired_at"),
    ),
    (
        "measured_by",
        ("measured_by", "measured_by_id", "operator_id", "operator", "sonographer", "entered_by"),
    ),
    (
        "clinical_notes",
        ("clinical_notes", "notes", "comment", "comments", "clinical_note"),
    ),
    ("given_name", ("given_name", "first_name", "patient_given_name")),
    ("family_name", ("family_name", "last_name", "surname", "patient_family_name")),
    ("date_of_birth", ("date_of_birth", "dob", "birth_date", "patient_dob")),
    ("gestational_age_weeks", ("gestational_age_weeks", "ga_weeks", "ga")),
):
    for a in aliases:
        _HEADER_TO_CANONICAL[_norm_header(a)] = canon


def _map_row(raw: dict[str, str]) -> dict[str, Any]:
    out: dict[str, Any] = {}
    for key, val in raw.items():
        if key is None or val is None:
            continue
        nk = _norm_header(str(key))
        canon = _HEADER_TO_CANONICAL.get(nk)
        if canon is None:
            continue
        out[canon] = val
    return out


def _parse_optional_float(s: str) -> float | None:
    t = s.strip()
    if t == "":
        return None
    return float(t)


def _parse_required_float(s: str | None, name: str) -> float:
    if s is None or str(s).strip() == "":
        raise ValueError(f"missing {name}")
    return float(str(s).strip())


def _parse_optional_date(s: str | None) -> date | None:
    if s is None:
        return None
    t = str(s).strip()
    if t == "":
        return None
    return date.fromisoformat(t[:10])


def _parse_measured_at(s: str | None) -> datetime:
    if s is None or str(s).strip() == "":
        raise ValueError("missing measured_at (scan / measurement time)")
    t = str(s).strip()
    if t.endswith("Z"):
        t = t[:-1] + "+00:00"
    if " " in t and "T" not in t:
        t = t.replace(" ", "T", 1)
    return datetime.fromisoformat(t)


def _row_to_measurement_payload(mapped: dict[str, Any]) -> dict[str, Any]:
    vi = _parse_required_float(mapped.get("vi_mm"), "vi_mm")
    ahw = _parse_required_float(mapped.get("ahw_mm"), "ahw_mm")
    tod = _parse_required_float(mapped.get("tod_mm"), "tod_mm")
    payload: dict[str, Any] = {"vi_mm": vi, "ahw_mm": ahw, "tod_mm": tod}
    if mapped.get("vi_percentile") not in (None, ""):
        payload["vi_percentile"] = _parse_optional_float(str(mapped["vi_percentile"]))
    if mapped.get("vi_p97_reference_mm") not in (None, ""):
        payload["vi_p97_reference_mm"] = _parse_optional_float(str(mapped["vi_p97_reference_mm"]))
    return payload


def _opt_str(mapped: dict[str, Any], key: str) -> str | None:
    v = mapped.get(key)
    if v is None or str(v).strip() == "":
        return None
    return str(v).strip()


def _row_to_patient_and_context(mapped: dict[str, Any]) -> tuple[PatientInfo, MeasurementContext]:
    ext = mapped.get("external_ref")
    if ext is None or str(ext).strip() == "":
        raise ValueError("missing patient external_ref (mrn / patient_id / external_ref)")
    ga = None
    if mapped.get("gestational_age_weeks") not in (None, ""):
        ga = _parse_optional_float(str(mapped["gestational_age_weeks"]))
    patient = PatientInfo(
        external_ref=str(ext).strip(),
        given_name=_opt_str(mapped, "given_name"),
        family_name=_opt_str(mapped, "family_name"),
        date_of_birth=_parse_optional_date(mapped.get("date_of_birth")),
        gestational_age_weeks=ga,
    )
    ctx = MeasurementContext(
        measured_at=_parse_measured_at(mapped.get("measured_at")),
        measured_by=_opt_str(mapped, "measured_by"),
        clinical_notes=_opt_str(mapped, "clinical_notes"),
    )
    return patient, ctx


def _cell_to_str(val: Any) -> str:
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.isoformat()
    if isinstance(val, date):
        return val.isoformat()
    return str(val).strip()


def _iter_csv_dicts(content: bytes) -> tuple[list[str], list[dict[str, str]]]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(StringIO(text))
    if reader.fieldnames is None:
        return [], []
    rows: list[dict[str, str]] = []
    for row in reader:
        rows.append({k: (v if v is not None else "") for k, v in row.items()})
    return [], rows


def _iter_xlsx_dicts(content: bytes) -> tuple[list[str], list[dict[str, str]]]:
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return [], []
    headers = [str(c).strip() if c is not None else "" for c in header_row]
    out: list[dict[str, str]] = []
    for tup in rows_iter:
        if tup is None or all(c is None or str(c).strip() == "" for c in tup):
            continue
        d: dict[str, str] = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            val = tup[i] if i < len(tup) else None
            d[h] = _cell_to_str(val)
        out.append(d)
    wb.close()
    return [], out


def import_measurements_from_tabular(content: bytes, filename: str) -> TabularImportResponse:
    """Import rows from ``.csv`` or ``.xlsx``. Each row needs patient id + ``measured_at`` + metrics."""
    name = (filename or "").lower()
    if name.endswith(".csv"):
        _, raw_rows = _iter_csv_dicts(content)
    elif name.endswith(".xlsx"):
        _, raw_rows = _iter_xlsx_dicts(content)
    elif name.endswith(".xls"):
        raise ValueError(
            "Legacy .xls is not supported; save as .xlsx or export to CSV."
        )
    else:
        raise ValueError(
            "Unsupported file type; use a .csv or .xlsx file (UTF-8 CSV recommended)."
        )

    ok: list[TabularImportRowOut] = []
    skipped: list[TabularImportSkippedRow] = []

    for i, raw in enumerate(raw_rows, start=2):
        mapped = _map_row(raw)
        try:
            patient, ctx = _row_to_patient_and_context(mapped)
            payload = _row_to_measurement_payload(mapped)
            m = VentricularMeasurements.model_validate(payload)
        except (ValueError, ValidationError) as e:
            skipped.append(TabularImportSkippedRow(row_number=i, reason=str(e)))
            continue
        ok.append(
            TabularImportRowOut(row_number=i, patient=patient, context=ctx, measurements=m)
        )

    return TabularImportResponse(rows=ok, skipped_rows=skipped)
